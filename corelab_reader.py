# -*- coding: utf-8 -*-
"""
Created on Mon Jun  8 19:26:05 2020

@author: Dmitry Molokhov (molokhov@outlook.com)

GitHub: https://github.com/dimmol/gamma_dist
"""

from openpyxl import load_workbook
import pandas as pd
import numpy as np
import re
import scipy.special as sps
import scipy.stats as stats
import scipy.optimize as optim
import time
import matplotlib.pyplot as plt

# Class to store sample data and relevant fluid properties.
class FlashExperimentData:
    
    def __init__(self, lqd=None, gas=None, res=None, ave_lqd_mw=None):
        self.liquid = lqd
        self.gas = gas
        self.reservoir = res
        self.av_lqd_mw = ave_lqd_mw
        # Estimation of average MW of C7 & C10 plus fractions
        # can be implemented later.
        self._ave_C10_mw = None
        self._ave_C7_mw = None
        self._c10_heavy_end_lqd = None
        self._c7_heavy_end_lqd = None
        self.gamma_input = None
        self.gamma_output = None
        self.depth = None
        self.cylinder = None
        self.pc_db = pd.read_excel(r'.\DATA\Components.xlsx', sheet_name = 'Sheet1',
                   nrows = 53, usecols = 'A:G, I', header = 0, 
                   na_values = [''])
    
    def assign_composition(self):
        pass
    
    # def _test(self, n=10):
    #     df = self.liquid[self.liquid['lqd_wp'] != 0].copy()
    #     df.reset_index(drop=True, inplace=True)
    #     i = df[df['scn'] == 'C'+str(n)].index[0]
    #     print(i)
    #     mw = self.av_lqd_mw
    #     df = df.merge(self.pc_db[['CoreLab Name', 'MW_lab']], how='left', 
    #                   left_on=['cl_name'], right_on=['CoreLab Name']).drop(['CoreLab Name'], axis=1)
    #     df.loc[i:, 'MW_lab'] = df.apply(lambda x : x['lqd_wp']*mw/x['lqd_mp'], axis = 1)
    #     print(df)
    
    # The following property decorators are to define the heavy 
    # end of the liquid fraction.
    # Don't like this way of splitting the heavy end but I reckon
    # I exhausted my knowledge of Pandas.
    @property
    def c10_heavy_end_lqd(self):
        i = self.liquid[self.liquid['scn'] == 'C10'].index[0]
        self._c10_heavy_end_lqd = self.liquid.iloc[i:]
        return self._c10_heavy_end_lqd
    
    # Calculating MW for the heavy end using lab MWs for light SCNs. These are often
    # inconsistent with book MWs for light components. This will introduce some errors in
    # resulting MW but this will hopefully be addressed by regression.
    def _calculate_MW(self, n=10):
        df = self.liquid[self.liquid['lqd_wp'] != 0].copy()
        df.reset_index(drop=True, inplace=True)
        mw = self.av_lqd_mw
        i = df[df['scn'] == 'C'+str(n)].index[0]
        # df['MW_lab'] = df.apply(lambda x : x['lqd_wp']*mw/x['lqd_mp'], axis = 1)
        df = df.merge(self.pc_db[['CoreLab Name', 'MW_lab']], how='left', 
                      left_on=['cl_name'], right_on=['CoreLab Name']).drop(['CoreLab Name'], axis=1)
        df.loc[i:, 'MW_lab'] = df.apply(lambda x : x['lqd_wp']*mw/x['lqd_mp'], axis = 1)
        a = df.loc[i:, 'lqd_wp'].sum()/100
        b = 1/self.av_lqd_mw
        c = df.loc[:i-1, 'lqd_wp'].div(df.loc[:i-1, 'MW_lab']).sum()/100
        return a/(b-c)

    @property
    def ave_C10_mw(self):
        self._ave_C10_mw = self._calculate_MW()
        return self._ave_C10_mw
    
    @property
    def c7_heavy_end_lqd(self):
        i = self.liquid[self.liquid['scn'] == 'C7'].index[0]
        self._c7_heavy_end_lqd = self.liquid.iloc[i:]
        return self._c7_heavy_end_lqd
    
    def prepare_input(self, id, n=10):
        lqd = self.c10_heavy_end_lqd
        mw = self.av_lqd_mw
        self.gamma_input = lqd.copy()
        self.gamma_input['MWi_lab'] = self.gamma_input.apply(lambda x : x['lqd_wp']*mw/x['lqd_mp'], axis = 1)
        # Calculating initial values of upper bounds of individual MWn slices.
        # Upper bounds are considered as midway between SCN MW values.
        # C36+ upper bound is set to an arbitrary number (10000).
        self.gamma_input['ubound_init'] = self.gamma_input['MWi_lab']+(self.gamma_input['MWi_lab'].shift(-1)-self.gamma_input['MWi_lab'])/2
        self.gamma_input['ubound_init'] = self.gamma_input['ubound_init'].fillna(100000)
        
        # Generating regresion variables for component molecular weight bounds:
        self.gamma_input['ubound'] = 'm'+self.gamma_input['scn']
        
        # Adding top row to represent a lower boundary of C10 (or upper C9 boundary)
        top_index = 'C'+str(n-1)
        df_top = pd.DataFrame(pd.DataFrame([[top_index]+[np.nan] * (len(self.gamma_input.columns)-1)], columns=self.gamma_input.columns))
        self.gamma_input = df_top.append(self.gamma_input, ignore_index=True)
        
        # Calculating rescaled C10+ weight fractions.
        self.gamma_input['wni_lab'] = self.gamma_input['lqd_wp']/self.gamma_input['lqd_wp'].sum()
        
        # Adding a regression variable for lower C10 molecular weight boundary:
        self.gamma_input.at[0,'ubound'] = 'ita'
        self.gamma_input.iloc[-1, self.gamma_input.columns.get_loc('ubound')] = self.gamma_input.iloc[-1, self.gamma_input.columns.get_loc('ubound_init')]
        self.gamma_input['sample_id'] = id.replace('.', '_')
        self.gamma_input['heavy_end_mw'] = id.replace('.', '_')+'_heavy_mw'


# Class to store multiple sample data.
class FlashExpDataCollection(FlashExperimentData, dict):
    
    def __init__(self, worksheet_names):
        dict.__init__(self, ((wn, FlashExperimentData.__init__(self)) for wn in worksheet_names))
        
    @property
    def sample_names(self):
        return list(self.keys())
    
    def add_sample(self, sample_name, sample):
        self[sample_name] = sample
        
    def _prepare_regression(self):
        for key, item in self.items():
            item.prepare_input(key)
            
    def gamma_distribution(self, reg_vals, reg_vars, rmse_switch = False):

        # Ensuring consistency of input data
        assert len(reg_vals) == len(reg_vars)
        
        # Creating a dictionary of regression values indexed with variable names:
        lookup = dict()
        for i in range(len(reg_vals)):
            lookup[reg_vars[i]] = reg_vals[i]
        
        # Updating the dataframe with set regression values (replacing variables with values):
        # Note, this time it is not a single dataframe but a class containing a number of dataframes.
        error_array = pd.Series(dtype='float64')
        for key, item in self.items():
            df = item.gamma_input
            df = df.replace(lookup)
            ind = key.replace('.', '_')+'_heavy_mw'
            # Below equation references are from the SPE Phase Behavior monograph.
            beta = (lookup[ind]-lookup['ita'])/lookup['alpha'] # Equation 5.14
            df['y'] = (df['ubound']-lookup['ita'])/beta # Equation 5.22
            # Below is the equation 5.20 from the monograph but with correction of the typo.
            # Correct form can be derived from the Equation 5.13.
            # Another typo is in the Equation 5.15 of the original SPE monograph. Correct
            # form of the equation can be found in 1990 Whiton's paper "Application of the Gamma
            # Distribution Model to MW and Boiling Point Data For Petroleum Fractions", Equation 21
            df['Q'] = (np.exp(-df['y'])*(df['y']**lookup['alpha'])/
                       sps.gamma(lookup['alpha']))
            df['P0'] = stats.gamma.cdf((df['ubound']-lookup['ita']), 
                                       a=lookup['alpha'], scale=beta) # Equation 5.18
            df['P1'] = df['P0']-(df['Q']/lookup['alpha']) # Equation 5.19
            df['Mi'] = (lookup['ita']+lookup['alpha']*beta*
                        ((df['P1']-df['P1'].shift())/(df['P0']-df['P0'].shift()))) # Equation 5.17
            df['Wi'] = df['Mi'] * (df['P0']-df['P0'].shift()) # Weight
            df['Wni'] = df['Wi']/df['Wi'].sum(skipna = True) # Normalised weight fraction
            
            temp = (df.loc[df.index[1:-1], 'Wni']-df.loc[df.index[1:-1], 'wni_lab'])**2
            error_array = pd.concat([error_array, temp], ignore_index=True)
            
            if rmse_switch:
                df['Zni'] = df['Wni']/df['Mi']*df['Wi'].sum(skipna = True)
                item.gamma_output = df
    
        return 100*error_array.mean()**0.5
        
    def gamma_distribution_fit(self, n=10, alpha=1):
        # Gamma distribution fit from C7 is not implemented yet.
        if n == 7:
            raise NotImplementedError
        self._prepare_regression()
        df = self[self.sample_names[0]].gamma_input
        reg_variables = np.concatenate((np.array(['alpha']), df.loc[df.index[0:-1], 'ubound'].unique(),
                                        np.array([sample.replace('.', '_')+'_heavy_mw' for sample in self.sample_names])))
        init_vals = df.loc[df.index[0:-1], 'ubound_init']
        init_vals.iloc[0] = init_vals.iloc[1]-14
        init_vals = pd.Series(alpha).append(init_vals, ignore_index=True)
        for key, item in self.items():
            init_vals = init_vals.append(pd.Series(item.ave_C10_mw), ignore_index=True)
            
        ub = init_vals+init_vals*0.02
        ub[17:] = init_vals[17:]+init_vals[17:]*0.05
        lb = init_vals-init_vals*0.02
        lb[17:] = init_vals[17:]-init_vals[17:]*0.05
        lb[0] = -np.inf
        ub[0] = np.inf
    
        res = optim.minimize(self.gamma_distribution, args=(reg_variables), x0=init_vals,
                             method = 'SLSQP', bounds=optim.Bounds(lb, ub), options={'maxiter':10000})
        res_df = pd.DataFrame({'Variables': reg_variables, 'Values':res.x})
        print('RMSE: ', res.fun)
        print(res_df)
        res_df = res_df.append({'Variables' : 'RMSE', 'Values' : res.fun}, ignore_index=True)
        res_df.to_csv(r'.\DATA\results.csv')

        self.gamma_distribution(res.x, reg_variables, rmse_switch = True)
        
    def gamma_distribution_export(self, file_path):
        
        for key, item in self.items():
            item.gamma_output.to_csv(file_path, mode='a') # [['SCN', 'Mi', 'Wni', 'Zni']]
            
    def _sample_plot(self, df, cylinder=None, depth=None):
        # Creating a plot of lab vs calculated compositions
        plt.style.use('classic')
        fig = plt.figure(figsize=[7,5])
        fig.suptitle('Laboratory vs Calculated Data')
        ax = plt.subplot(111)
        ax.set_xlabel('Calculated Molecular Weight, g/mol')
        ax.set_ylabel('Normalized Weight Fractions')
        ax.set_title('Cylinder: '+str(cylinder)+' Depth: '+str(depth)+'m')
        ax.grid('on')
        ax.set_yscale('log')
        ax.set_xlim(0, 700)
        ax.set_ylim(0.001, 1)
        ax.xaxis.set_tick_params(size=0)
        xlab = ax.xaxis.get_label()
        ylab = ax.yaxis.get_label()
        xlab.set_style('italic')
        xlab.set_size(10)
        ylab.set_style('italic')
        ylab.set_size(10)
        ttl = ax.title
        ttl.set_weight('bold')
        ax.plot(df['Mi'], df['wni_lab'], '-ok', markerfacecolor='w', label='Laboratory')
        ax.plot(df['Mi'], df['Wni'], 'or', label='Calculated')
        ax.legend(loc='best', frameon=True, fontsize=10)
        plt.show()
            
    def gamma_distribution_plot(self):
        for key, item in self.items():
            self._sample_plot(item.gamma_output, item.cylinder, item.depth)
        
# Class that contains functionality to parse a Core Labs Excel report
# and pass the data to FlashExperimentData class instance.
# openpyxl only works with Excel 2010+ file format.
class CoreLabsXLSXLoader:
    
    # Depending whether a specific worksheet is input a class instance has
    # a functionality either to read the whole flash data found in the
    # workbook or just that individual worksheet.
    def __init__(self, report_path, worksheet=None):
        self.file_path = report_path
        self.worksheet = worksheet
    
    # Method to parse an individual worksheet.
    def read_flash_data(self, worksheet):
        # Reading the composition table as the whole.
        df = pd.read_excel(self.file_path, worksheet, skiprows = 11, nrows = 52, 
                        usecols = 'B:I', header = None,
                        names = ['scn', 'cl_name', 'lqd_mp', 'lqd_wp', 'gas_mp',
                                 'gas_wp', 'res_mp', 'res_wp'], na_values = ' ')
        # Filling in empty cells in carbon group columns.
        df['scn'] = df['scn'].ffill()
        df.set_index(['scn', 'cl_name'], inplace=True)
        # df.sort_index(inplace=True)
        liq = df.iloc[:, 0:2].reset_index()
        gas = df.iloc[:, 2:4].reset_index()
        res = df.iloc[:, 4:].reset_index()
        return liq, gas, res
    
    def read(self):
        wb = load_workbook(self.file_path)
        if self.worksheet:
            flash_data_list = self.worksheet
        else:
            flash_data_list = [worksheet for worksheet in wb.sheetnames if
                               re.search('C\.\d+', worksheet)]
        samples = FlashExpDataCollection(flash_data_list)
        for worksheet in flash_data_list:
            sheet = wb[worksheet]
            desc = sheet['B8'].value+sheet['B9'].value
            depth, sample_num, cylinder = self.__parser(desc)
            # print('Depth: ', depth, 'Sample number: ', sample_num, 'Cylinder: ', cylinder)
            liq, gas, res = self.read_flash_data(worksheet)
            # Typically, flashed liquid average mole weight in CL reports is in
            # the cell O39:
            lqd_av_mw = sheet['O39'].value
            samples.add_sample(worksheet, FlashExperimentData(liq, gas, res,
                                                              lqd_av_mw))
            samples[worksheet].depth = depth
            samples[worksheet].cylinder = cylinder
        return samples
    
    # Parcer function is supposed to extract useful sample descriptors from
    # the text strings above the composition table. In most Core LAbs reports
    # these strings are in the cells B8 and B9. Presumably, they contain
    # sample depth, cylinder and sample numbers. At this stage I am only
    # extracting numerical part of the these numbers. Depth is assumed to be
    # MD depth in metres.
    def __parser(self, descript_string):
        try:
            depth = float(re.search('Depth\D+(\d+\.?\d*)', descript_string).group(1))
        except:
            depth = None
        try:
            sample_num = re.search('Sample N\D+(\d+\.?\d*)', descript_string).group(1)
        except:
            sample_num = 'N/A'
        try:
            cylinder = re.search('(Cylinder|Chamber)\D+(\d+)', descript_string).group(2)
        except:
            cylinder = 'N/A'
        return depth, sample_num, cylinder

if __name__ == "__main__":
    
    start_time = time.time()
    input_file = '.\DATA\PS1.xlsx'
    cl_report = CoreLabsXLSXLoader(input_file) #, worksheet=['C.1', 'C.4']
    sample_collection = cl_report.read()
    sample_collection.gamma_distribution_fit()
    sample_collection.gamma_distribution_export(r'.\DATA\gamma.csv')
    sample_collection.gamma_distribution_plot()
    
    print("--- Execution time %s seconds ---" % (time.time() - start_time))